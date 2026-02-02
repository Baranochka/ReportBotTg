import database as db
from datetime import date
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


class ReportExcel:
    def __init__(self):
        name_file = None
        self.create_excel_file()

    def create_excel_file(self):
        # Создаем новую книгу
        wb = Workbook()

        fill_yellow = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid") # Желтый цвет
        fill_yellow2 = PatternFill(start_color="fffc00", end_color="fffc00", fill_type="solid") # Светло-Желтый цвет
        fill_green = PatternFill(start_color="76933c", end_color="76933c", fill_type="solid")   # Зеленый цвет
        thin = Side(border_style="thin", color="000000")   # тонкая чёрная линия
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        # Записываем данные
        sheet = wb.active
        
        for i in range(1,35):
            sheet.cell(row=1, column=i).fill = fill_yellow
            sheet.cell(row=1, column=i).border = border
            sheet.cell(row=2, column=i).fill = fill_yellow
            sheet.cell(row=2, column=i).border = border
            sheet.cell(row=3, column=i).fill = fill_green
            sheet.cell(row=3, column=i).border = border
        
        sheet.cell(row=1, column=1, value="ФИ")

        
        sheet.cell(row=1, column=8, value="1 неделя")
        sheet.cell(row=1, column=14, value="2 неделя")
        sheet.cell(row=1, column=20, value="3 неделя")
        sheet.cell(row=1, column=26, value="4 неделя")
        sheet.cell(row=1, column=32, value="5 неделя")
        sheet.cell(row=1, column=33, value="Факт")
        sheet.cell(row=1, column=34, value="План")

        data = []
        for i in range(1,6):
            print(f'week_{i}')
            buf = db.select_all(f"week_{i}")
            data.append(buf)
        
        row = 4
        
        buf = db.select_all("list_date")
        for index_week in range(0,5):
            sheet.cell(row=2, column=3+index_week*6, value=buf[0][index_week*7+1])
            sheet.cell(row=2, column=4+index_week*6, value=buf[0][index_week*7+2])
            sheet.cell(row=2, column=5+index_week*6, value=buf[0][index_week*7+3])
            sheet.cell(row=2, column=6+index_week*6, value=buf[0][index_week*7+4])
            sheet.cell(row=2, column=7+index_week*6, value=buf[0][index_week*7+5])
        
        
        sheet.cell(row=3, column=2, value="Исходящие звонки")
        row = repeat(sheet, row, data, 0,   "Расклейка, рассылка, раскидка")
        row = repeat(sheet, row, data, 10,  "Встречи")
        row = repeat(sheet, row, data, 20,  "Сторис,пост")
        row = repeat(sheet, row, data, 30,  "Агентский договор")
        row = repeat(sheet, row, data, 40,  "Задаток/Бронь")
        row = repeat(sheet, row, data, 50,  "Закрытые сделки")
        row = repeat(sheet, row, data, 60,  "Расторжение АД")
        row = repeat(sheet, row, data, 70,  "Входящий по объекту")
        row = repeat(sheet, row, data, 80,  "Показ (продажа)")
        row = repeat(sheet, row, data, 90,  "Показ (подбор)")
        row = repeat(sheet, row, data, 100, "Отчёт собственнику")
        row = repeat(sheet, row, data, 110, "")

        sheet.column_dimensions.group(start="C", end="G", hidden=True)
        sheet.column_dimensions.group(start="I", end="M", hidden=True)
        sheet.column_dimensions.group(start="O", end="S", hidden=True)
        sheet.column_dimensions.group(start="U", end="Y", hidden=True)
        sheet.column_dimensions.group(start="AA", end="AE", hidden=True)
        sheet.column_dimensions["A"].width = 25   # ширина колонки A
        sheet.column_dimensions["B"].width = 30   # ширина колонки B
        self.name_file = f"Отчёт-на-{str(date.today())}.xlsx"
        # Сохраняем файл
        wb.save(f"shared/{self.name_file}")

def repeat(sheet, row, data, val, title):
    fill_yellow2 = PatternFill(start_color="fffc00", end_color="fffc00", fill_type="solid") # Светло-Желтый цвет
    fill_green = PatternFill(start_color="76933c", end_color="76933c", fill_type="solid")   # Зеленый цвет
    thin = Side(border_style="thin", color="000000")   # тонкая чёрная линия
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for index_people in range(len(data[0])): # Проход по людям (строкам)
        for index_week in range(0,5):
            sheet.cell(row=row, column=1, value=data[index_week][index_people][1])
            
            sheet.cell(row=row, column=3+index_week*6, value=data[index_week][index_people][2+val])
            if data[index_week][index_people][3+val] != None:
                cell = sheet.cell(row=row, column=3+index_week*6)
                cell.comment = Comment(data[index_week][index_people][3+val], "Автор")
            
            sheet.cell(row=row, column=4+index_week*6, value=data[index_week][index_people][4+val])
            if data[index_week][index_people][5+val] != None:
                cell = sheet.cell(row=row, column=4+index_week*6)
                cell.comment = Comment(data[index_week][index_people][5+val], "Автор")
            
            sheet.cell(row=row, column=5+index_week*6, value=data[index_week][index_people][6+val])
            if data[index_week][index_people][7+val] != None:
                cell = sheet.cell(row=row, column=5+index_week*6)
                cell.comment = Comment(data[index_week][index_people][7+val], "Автор")
            
            sheet.cell(row=row, column=6+index_week*6, value=data[index_week][index_people][8+val])
            if data[index_week][index_people][9+val] != None:
                cell = sheet.cell(row=row, column=6+index_week*6)
                cell.comment = Comment(data[index_week][index_people][9+val], "Автор")
            
            sheet.cell(row=row, column=7+index_week*6, value=data[index_week][index_people][10+val])
            if data[index_week][index_people][11+val] != None:
                cell = sheet.cell(row=row, column=7+index_week*6)
                cell.comment = Comment(data[index_week][index_people][11+val], "Автор")
            
        sheet.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")
        sheet.cell(row=row, column=14, value=f"=SUM(I{row}:M{row})")
        sheet.cell(row=row, column=20, value=f"=SUM(O{row}:S{row})")
        sheet.cell(row=row, column=26, value=f"=SUM(U{row}:Y{row})")
        sheet.cell(row=row, column=32, value=f"=SUM(AA{row}:AE{row})")
        sheet.cell(row=row, column=33, value=f"=SUM(H{row},N{row},T{row},Z{row},AF{row})")
        for i in range(1,35):
            sheet.cell(row=row, column=i).border = border
        row +=1
    sheet.cell(row=row, column=1, value="Итого")
    for i in range(1,35):
        sheet.cell(row=row, column=i).fill = fill_yellow2
        sheet.cell(row=row, column=i).border = border
    sheet.cell(row=row, column=8, value=f"=SUM(H{row-len(data[0])}:H{row-1})")
    sheet.cell(row=row, column=14, value=f"=SUM(N{row-len(data[0])}:N{row-1})")
    sheet.cell(row=row, column=20, value=f"=SUM(T{row-len(data[0])}:T{row-1})")
    sheet.cell(row=row, column=26, value=f"=SUM(Z{row-len(data[0])}:Z{row-1})")
    sheet.cell(row=row, column=32, value=f"=SUM(AF{row-len(data[0])}:AF{row-1})")
    sheet.cell(row=row, column=33, value=f"=SUM(H{row},N{row},T{row},Z{row},AF{row})")
    row +=2
    if title != "":
        sheet.cell(row=row, column=2, value=title)
        for i in range(1,35):
            sheet.cell(row=row, column=i).fill = fill_green
            sheet.cell(row=row, column=i).border = border
            sheet.cell(row=row-1, column=i).border = border
    row +=1
    return row
        